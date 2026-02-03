import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def parse_text_to_dataframe(text,
                            timestamp_pattern=r'^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}',
                            skip_patterns=None,
                            delimiter=None,
                            auto_detect=True):
    if skip_patterns is None:
        skip_patterns = []

    lines = text.strip().split('\n')
    headers = []
    data_rows = []

    for line in lines:

        if timestamp_pattern:
            line = re.sub(timestamp_pattern, '', line).strip()

        if not line:
            continue

        should_skip = False
        for pattern in skip_patterns:
            if re.search(pattern, line, re.IGNORECASE):
                should_skip = True
                break
        if should_skip:
            continue

        if delimiter:
            parts = [p.strip() for p in line.split(delimiter)]
        else:
            parts = line.split()

        if not headers and auto_detect:
            # 检查是否包含非数字列名
            if any(not is_number(p) for p in parts):
                headers = parts
                continue

        processed_parts = []
        for p in parts:
            if is_number(p):
                try:
                    if '.' in p:
                        processed_parts.append(float(p))
                    else:
                        processed_parts.append(int(p))
                except:
                    processed_parts.append(p)
            else:
                processed_parts.append(p)

        data_rows.append(processed_parts)

    if not headers and data_rows:
        max_cols = max(len(row) for row in data_rows)
        headers = [f'Column_{i + 1}' for i in range(max_cols)]

    max_cols = len(headers)
    for i, row in enumerate(data_rows):
        if len(row) < max_cols:
            data_rows[i] = row + [None] * (max_cols - len(row))
        elif len(row) > max_cols:
            data_rows[i] = row[:max_cols]

    df = pd.DataFrame(data_rows, columns=headers)
    return df


def is_number(s):
    try:
        float(s)
        return True
    except:
        return False


def create_excel(df, output_path,
                 title=None,
                 header_bg_color='4472C4',
                 header_font_color='FFFFFF',
                 freeze_panes=True):
    # 保存为Excel
    df.to_excel(output_path, index=False, sheet_name=title or 'Sheet1')

    # 加载并格式化
    wb = load_workbook(output_path)
    sheet = wb.active

    # 设置表头格式
    header_fill = PatternFill(start_color=header_bg_color,
                              end_color=header_bg_color,
                              fill_type='solid')
    header_font = Font(bold=True, color=header_font_color, name='Arial', size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据格式
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(vertical='center')

    # 自动调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    if freeze_panes:
        sheet.freeze_panes = 'A2'

    wb.save(output_path)
    return output_path


def text_to_excel(text, output_path, **kwargs):
    # 分离参数
    parse_params = ['timestamp_pattern', 'skip_patterns', 'delimiter', 'auto_detect']
    excel_params = ['title', 'header_bg_color', 'header_font_color', 'freeze_panes']

    parse_kwargs = {k: v for k, v in kwargs.items() if k in parse_params}
    excel_kwargs = {k: v for k, v in kwargs.items() if k in excel_params}

    df = parse_text_to_dataframe(text, **parse_kwargs)
    create_excel(df, output_path, **excel_kwargs)

    return df


# 使用示例
if __name__ == '__main__':
    sample_text = """ 2026-02-01 11:15:25
Layers sorted by improvement (higher = more beneficial):
2026-02-01 11:15:25
 layer_name  ssim_score  final_improvement
2026-02-01 11:15:25
features.17   -0.510974               0.26
2026-02-01 11:15:25
features.20   -0.526358               0.14
2026-02-01 11:15:25
features.10    0.847710               0.06
2026-02-01 11:15:25
features.14   -0.129671               0.02
2026-02-01 11:15:25
features.34    0.500000              -0.07
2026-02-01 11:15:25
features.37    0.500000              -0.09
2026-02-01 11:15:25
 features.0    0.963572              -0.10
2026-02-01 11:15:25
 classifier    0.500000              -0.12
2026-02-01 11:15:25
features.30    0.979052              -0.15
2026-02-01 11:15:25
features.40    0.500000              -0.16
2026-02-01 11:15:25
 features.7    0.967294              -0.21
2026-02-01 11:15:25
 features.3    0.961631              -0.24
2026-02-01 11:15:25
features.27    0.838445              -0.33
2026-02-01 11:15:25
features.24   -0.777554              -0.44"""

    df = text_to_excel(
        sample_text,
        'example_output.xlsx',
        skip_patterns=[r'sorted by', r'beneficial'],
        title='Layer Analysis'
    )


